# -*- coding: utf-8 -*-
"""
开发者：周梦雄
最后更新日期：2020/5/18
"""
import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QMainWindow,
    QTableWidget,
    QMessageBox,
)
from Ui_chip_id_assignment import *
from openpyxl import load_workbook
import time
import subprocess
# import pyautogui
import win32api
import win32gui
import win32con
import win32clipboard
import time
import win32com.client
from PyQt5.QtGui import QIntValidator, QRegExpValidator
from PyQt5.QtCore import QRegExp


class MyMainWindow(QMainWindow, Ui_chip_id_assignment):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # 整型，范围：[1, 99]
        # qtyIntValidator = QIntValidator(1, 99)
        # self.le_qty.setValidator(qtyIntValidator)
        reg2 = QRegExp("^[1-9][0-9]$")
        qtyValidator = QRegExpValidator(reg2)
        self.le_qty.setValidator(qtyValidator)
        # 审批编号6位数字
        reg = QRegExp("[0-9]+$")
        approvalValidator = QRegExpValidator(reg)
        self.le_approval.setValidator(approvalValidator)
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")        
        self.btn_assign.clicked.connect(self.id_assign)
        self.le_approval.editingFinished.connect(self.approval_check)
        self.show()

    # 检查是否输入6位审批单号
    def approval_check(self):
        if len(self.le_approval.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval.setFocus()

    # 分配ID方法
    def id_assign(self):
        if len(self.le_approval.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval.setFocus()
            return
        if len(self.le_qty.text()) == 0:
            QMessageBox.warning(
            self, '错误：', '您未输入数量，请输入！', QMessageBox.Ok)
            self.le_qty.setFocus()
            return
        QMessageBox.warning(
            self, '请注意：', '点击OK后，双手离开鼠标和键盘，等待ID自动分配完毕！', QMessageBox.Ok)
        if self.cb_type_id.currentText() == "智芯国网送检ID":
            #  ID数据库路径-----------------（修改为实际生产路径）
            path = r"C:\Users\zhoum\Desktop\20200408\20190408.xlsx"
        elif self.cb_type_id.currentText() == "智芯ID":
            path = r"C:\Users\Administrator\Desktop\小单ID\小单ID第三批1000\510079-511078 小单专用1000.xlsx"
        else:
            path = r"C:\Users\Administrator\Desktop\Chip_ID_Assign\4001-5000给周工1000个用于小批量订单专用.xlsx"
        # 打开excel工作簿
        wb = load_workbook(path)
        # 打开ID工作表（注意先将工作表名改为ID再运行程序）
        ws = wb['ID']
        # 输入派工单号
        wo = self.le_approval.text()
        # 输入需要ID的数量
        id_qty = self.le_qty.text()
        # 该ID使用在哪种产品
        pd = self.cb_type_prod.currentText()
        # 起始索引
        index = ws.cell(row=1, column=5).value
        # 结束索引
        end = int(index)+int(id_qty)
        # 判断ID数量是否充足
        if end > 1000:
            print("ID数量不足，请重新申请ID！！！")
            QMessageBox.warning(
            self, '错误：', 'ID数量不足，请重新申请ID！！！', QMessageBox.Ok)
            return
        else:
            for n in range(int(index), end):
                ws.cell(row=n, column=4).value = wo+pd
            # 打印首尾ID
            print(ws.cell(row=int(index), column=1).value)
            print(ws.cell(row=end-1, column=1).value)
            #  提示勿动鼠标
            # pyautogui.alert("请点击确定后，勿动鼠标键盘，电脑自动分ID并自动命名ID文件！！！")
            # 组合ID文件名
            id_name = str(ws.cell(row=int(index), column=3).value)+'-'+str(ws.cell(
                row=end-1, column=3).value)+'-'+str.upper(wo)+'-'+str.upper(pd)+'-'+id_qty+'pcs'
            id_name.upper()
            # print(id_name)
            if self.cb_type_id.currentText() == "智芯国网送检ID":
                # 电脑自动运行导出ID工具-----------------(修改为实际生产路径)
                win32api.ShellExecute(
                    0, 'open', r"C:\Users\zhoum\Desktop\20200408\FChipIdManger.exe", '', '', 1)
            elif self.cb_type_id.currentText() == "智芯ID":
                win32api.ShellExecute(
                    0, 'open', r"C:\Users\Administrator\Desktop\小单ID\小单ID第三批1000\FChipIdManger.exe", '', '', 1)                
            else:
                win32api.ShellExecute(
                    0, 'open', r"C:\Users\Administrator\Desktop\Chip_ID_Assign\FChipIdManger.exe", '', '', 1)                
            # time.sleep(3)
            # subprocess.Popen(r"C:\Users\Administrator\Desktop\小单ID\小单ID第二批1000\FChipIdManger.exe",stdout=subprocess.PIPE)
            # 【使用pyautogui方案】
            # # 电脑自动选中文本区域
            # pyautogui.click(990, 434, duration=0.2)
            # # 电脑自动输入ID数量
            # pyautogui.typewrite(id_qty)
            # # 电脑自动点击导出按钮
            # pyautogui.click(961, 597, duration=0.2)
            # # 暂停1s
            # time.sleep(1)
            # # 电脑自动选中ID文件名区域
            # pyautogui.click(900, 642, duration=0.2)
            # # 电脑自动输入ID文件名
            # pyautogui.typewrite(id_name, interval=0.2)
            # # pyautogui.press('enter')
            # # time.sleep(3)
            # # 电脑自动点击保存
            # pyautogui.click(1279, 732, duration=0.5)
            # # 电脑自动点击确定
            # pyautogui.click(999, 580, duration=0.5)
            # # 电脑自动关掉导ID工具
            # pyautogui.click(1092, 358, duration=0.5)
            # 【使用pywin32方案】
            time.sleep(0.5)
            try:
                # 获取FChipIdManger软件主窗口句柄
                hdl_main = win32gui.FindWindow(None, 'FChipIdManger')
                # print("主窗口句柄为：%x" % hdl_main)
                # 在主窗口查到数量文本框句柄
                hdl_edit = win32gui.FindWindowEx(hdl_main, None, 'Edit', None)
                # print("文本框句柄为：%x" % hdl_edit)
                # 在主窗口找到导出芯片ID文件按钮句柄
                hdl_btn1 = win32gui.FindWindowEx(
                    hdl_main, None, 'Button', None)
                # print("导出芯片ID文件按钮句柄为：%x" % hdl_btn1)
                # hdl_btn2 = win32gui.FindWindowEx(hdl_main, hdl_btn1, 'Button', None)
                # print("第二个按钮句柄为：%x" %hdl_btn2)
                # print("导出芯片ID文件按钮名称为：%s" % win32gui.GetWindowText(hdl_btn1))
                # print("第二个按钮名称为：%s" %win32gui.GetWindowText(hdl_btn2))
                # 使用win32gui.SendMessage发送数量到文本框
                win32gui.SendMessage(
                    hdl_edit, win32con.WM_SETTEXT, None, id_qty)
                # time.sleep(0.2)
                # 点击导出芯片ID文件按钮打开对话框
                win32gui.PostMessage(
                    hdl_btn1, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, 0)
                win32gui.PostMessage(
                    hdl_btn1, win32con.WM_LBUTTONUP, win32con.MK_LBUTTON, 0)
                # 根据电脑响应速度可增加或减少，默认2
                time.sleep(1)
                # 获取选择芯片ID文件对话窗口句柄
                hdl_select = win32gui.FindWindow(None, '选择芯片ID文件')
                # 获取保存按钮句柄
                hdl_btn3 = win32gui.FindWindowEx(
                    hdl_select, None, 'button', None)
                # print("保存按钮句柄为：%x" % hdl_btn3)
                # 获取选择芯片ID文件对话框窗口子窗口
                hwndChildList = []
                win32gui.EnumChildWindows(
                    hdl_select, lambda hwnd, param: param.append(hwnd),  hwndChildList)
                # print(hwndChildList)
                # ID文件保存路径
                filePath = r"C:\Users\zhoum\Desktop\20200408"
                shell = win32com.client.Dispatch("WScript.Shell")
                shell.SendKeys('%')
                # 将窗口设为系统的前台窗口
                win32gui.SetForegroundWindow(hdl_select)
                # 获取保存按钮坐标并左击
                x1, y1, x2, y2 = win32gui.GetWindowRect(hwndChildList[-9])
                win32api.SetCursorPos([x1, y1])
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP |
                                     win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                # 将路径复制到剪切板
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(filePath)
                win32clipboard.CloseClipboard()
                # 按下ctrl+v
                win32api.keybd_event(0x11, 0, 0, 0)
                win32api.keybd_event(0x56, 0, 0, 0)
                win32api.keybd_event(0x56, 0, win32con.KEYEVENTF_KEYUP, 0)
                win32api.keybd_event(0x11, 0, win32con.KEYEVENTF_KEYUP, 0)
                # 按回车进入该路径
                win32api.keybd_event(0x0D, 0, 0, 0)
                time.sleep(0.2)
                # 发送文件名
                win32gui.SendMessage(
                    hwndChildList[4], win32con.WM_SETTEXT, None, id_name)
                # # print(id_name)
                # 根据保存按钮句柄点击该按钮
                win32gui.PostMessage(
                    hdl_btn3, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, 0)
                win32gui.PostMessage(
                    hdl_btn3, win32con.WM_LBUTTONUP, win32con.MK_LBUTTON, 0)
                # 发送回车键
                win32api.keybd_event(13, 0, 0, 0)
                win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)
                # # 将窗口设为系统的前台窗口
                # shell = win32com.client.Dispatch("WScript.Shell")
                # shell.SendKeys('%')
                # win32gui.SetForegroundWindow(hdl_main)
                time.sleep(0.3)
                # 关闭FChipIdManger软件窗口
                win32gui.PostMessage(hdl_main, win32con.WM_CLOSE, 0, 0)
                self.statusbar.showMessage(id_name+"芯片ID分配成功！", 3000)
            except Exception:
                QMessageBox.warning(
                    self, '错误：', 'ID分配失败，请重新分配！', QMessageBox.Ok)
                return
            if win32gui.IsWindow(hdl_select):
                QMessageBox.warning(
                    self, '错误：', 'ID分配失败，请关掉除主窗口以外的窗口重新分配！', QMessageBox.Ok)
                return                                 
            # 更新索引
            ws.cell(row=1, column=5).value = end
            # 保存数据库
            wb.save(path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = MyMainWindow()
    sys.exit(app.exec_())
